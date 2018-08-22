from openerp import models, fields, api, _
import base64
import xlrd,datetime
from openerp.exceptions import Warning

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import cStringIO

class export_wsp_xls(models.TransientModel):
    _name = 'export.wsp.xls'
    _description = 'Exporting WSP from XLS files'
    
    @api.multi
    def get_population_group(self, race_value):
        race = ''
        if race_value == 'african':
            race = 'African'
        if race_value == 'coloured':
            race = 'Coloured'
        if race_value == 'indian':
            race = 'Indian'
        if race_value == 'white':
            race = 'White'
        return race
    
    @api.multi
    def get_genders(self, gender_value):
        gender = ''
        if gender_value == 'male':
            gender = 'M - Male'
        if gender_value == 'female':
            gender = 'F - Female'
        return gender
    
    @api.multi
    def get_disability_status(self, disability_value):
        disability_status = ''
        if disability_value == 'site':
            disability_status = '01 - Sight ( even with glasses )'
        if disability_value == 'hearing':
            disability_status = '02 - Hearing ( even with h.aid )'
        if disability_value == 'communication':
            disability_status = '03 - Communication ( talk/listen)'
        if disability_value == 'physical':
            disability_status = '04 - Physical ( move/stand, etc)'
        if disability_value == 'intellectual':
            disability_status = '05 - Intellectual ( learn,etc)'
        if disability_value == 'emotional':
            disability_status = '06 - Emotional ( behav/psych)'
        if disability_value == 'multiple':
            disability_status = '07 - Multiple'
        if disability_value == 'disabled':
            disability_status = '08 - Disabled'
        if disability_value == 'N-None':
            disability_status = 'none'
        return disability_status
    
    @api.multi
    def get_urban_rural(self, urban_value):
        urban_rural = ''
        if urban_value == 'urban' :
            urban_rural = 'Urban'
        if urban_value == 'rural' :
            urban_rural = 'Rural'
        if urban_value == 'unknown' :
            urban_rural = 'Unknown'
        return urban_rural
    
    @api.multi
    def get_aet_level(self, aet_level_value):
        aet_level = ''
        if aet_level_value == 'aet_level_1':
            aet_level = 'AET Level 1'
        if aet_level_value == 'aet_level_2':
            aet_level = 'AET Level 2'
        if aet_level_value == 'aet_level_3':
            aet_level = 'AET Level 3'
        if aet_level_value == 'aet_level_4':
            aet_level = 'AET Level 4'
        return aet_level
    
    @api.multi
    def get_aet_subject(self, aet_subject_value):
        aet_subject = ''
        if aet_subject_value == 'life_skills':
            aet_subject = 'Life Skills'
        if aet_subject_value == 'numeracy':
            aet_subject = 'Numeracy'
        if aet_subject_value == 'literacy':
            aet_subject = 'Literacy'
        return aet_subject
    
    @api.multi
    def get_training_type(self, training_type_value):
        training_type = ''
        if training_type_value == 'pivotal':
            training_type = 'Pivotal'
        if training_type_value == 'non-pivotal':
            training_type = 'Non Pivotal'
        return training_type 
    
    @api.multi
    def get_socio_eco_status(self, socio_value):
        socio = ''
        if socio_value == 'employed' :
            socio = 'Employed'
        if socio_value == 'unemployed':
            socio = 'Unemployed'
        return socio
    
    @api.multi
    def get_nqf_aligned(self, nqf_aligned_value):
        nqf_aligned = ''
        if nqf_aligned_value == 'yes':
            nqf_aligned = 'Yes'
        if nqf_aligned_value == 'no':
            nqf_aligned = 'No'
        return nqf_aligned
    
    @api.multi
    def get_nqf_level(self, nqf_level_value):
        nqf_level = ''
        if nqf_level_value == 'abet':
            nqf_level = 'ABET'
        if nqf_level_value == '':
            nqf_level = 'Level 1'
        if nqf_level_value == 'level2':
            nqf_level = 'Level 2'
        if nqf_level_value == 'level3':
            nqf_level = 'Level 3'
        if nqf_level_value == 'level4':
            nqf_level = 'Level 4'
        if nqf_level_value == 'level5':
            nqf_level = 'Level 5'
        if nqf_level_value == 'level6':
            nqf_level = 'Level 6'
        if nqf_level_value == 'level7':
            nqf_level = 'Level 7'
        if nqf_level_value == 'level8':
            nqf_level = 'Level 8'
        if nqf_level_value == 'level9':
            nqf_level = 'Level 9'
        if nqf_level_value == 'level10':
            nqf_level = 'Level 10'
        return nqf_level
    
    @api.multi
    def get_citizen_status(self, citizen_value):
        citizen_status = ''
        if citizen_value == 'dual' :
            citizen_status = 'D - Dual (SA plus other)'
        if citizen_value == 'other' :
            citizen_status = 'O - Other'
        if citizen_value == 'sa' :
            citizen_status = 'SA - South Africa'
        if citizen_value == 'unknown' :
            citizen_status = 'U - Unknown'
        return citizen_status
    
    @api.multi
    def get_id_type(self, id_value):
        id_status = ''
        if id_value == 'id_document' :
            id_status = 'ID Document'
        if id_value == 'passport' :
            id_status = 'Passport'
        return id_status
    
    @api.multi
    def get_highest_edu(self, high_edu_value):
        highest_edu = ''
        if high_edu_value == 'abet_level_1':
            highest_edu = 'Abet Level 1'
        if high_edu_value == 'abet_level_2':
            highest_edu = 'Abet Level 2'
        if high_edu_value == 'abet_level_3':
            highest_edu = 'Abet Level 3'
        if high_edu_value == 'abet_level_4':
            highest_edu = 'Abet Level 4'
        if high_edu_value == 'nqf123':
            highest_edu = 'NQF 1,2,3'
        if high_edu_value == 'nqf45':
            highest_edu = 'NQF 4,5'
        if high_edu_value == 'nqf67':
            highest_edu = 'NQF 6,7'
        if high_edu_value == 'nqf8910':
            highest_edu = 'NQF 8,9,10'
        return highest_edu
     

    @api.multi
    def export_previous(self):
        ''' This Method will export wsp template. '''
        workbook=Workbook()
        buffer = cStringIO.StringIO()
        context = self._context
        wsp_plan_data = self.env['wsp.plan'].browse(context.get('active_id',False))
        
        if wsp_plan_data.training_actual_ids :
            row=col=1
            worksheet1=workbook.active
            worksheet1.title = "Actual Training"
            
            h1=worksheet1.cell(column=col, row=row, value="Type of Training")
            worksheet1.merge_cells(start_row=row,start_column=col,end_row=row+1,end_column=col)
            h1.alignment=Alignment(horizontal='center',vertical='center')
            h1.font=Font(size=12,bold=True)
            col+=1
            
            h2=worksheet1.cell(column=col, row=row, value="Occupation")
            worksheet1.merge_cells(start_row=row,start_column=col,end_row=row+1,end_column=col)
            h2.alignment=Alignment(horizontal='center',vertical='center')
            h2.font=Font(size=12,bold=True)
            col+=1
            
            h3=worksheet1.cell(column=col, row=row, value="Specialisation")
            worksheet1.merge_cells(start_row=row,start_column=col,end_row=row+1,end_column=col)
            h3.alignment=Alignment(horizontal='center',vertical='center')
            h3.font=Font(size=12,bold=True)
            col+=1
            
            h4=worksheet1.cell(column=col, row=row, value="Municipality")
            worksheet1.merge_cells(start_row=row,start_column=col,end_row=row+1,end_column=col)
            h4.alignment=Alignment(horizontal='center',vertical='center')
            h4.font=Font(size=12,bold=True)
            col+=1
            
            h5=worksheet1.cell(column=col, row=row, value="Urban/Rural")
            worksheet1.merge_cells(start_row=row,start_column=col,end_row=row+1,end_column=col)
            h5.alignment=Alignment(horizontal='center',vertical='center')
            h5.font=Font(size=12,bold=True)
            col+=1
            
            h6=worksheet1.cell(column=col, row=row, value="Province")
            worksheet1.merge_cells(start_row=row,start_column=col,end_row=row+1,end_column=col)
            h6.alignment=Alignment(horizontal='center',vertical='center')
            h6.font=Font(size=12,bold=True)
            col+=1
            
            h7=worksheet1.cell(column=col, row=row, value="Socio Economic\nstatus")
            worksheet1.merge_cells(start_row=row,start_column=col,end_row=row+1,end_column=col)
            h7.alignment=Alignment(horizontal='center',vertical='center')
            h7.font=Font(size=12,bold=True)
            col+=1
            
            h8=worksheet1.cell(column=col, row=row, value="Type of Training\nIntervention/Pivotal\nProgramme Type")
            worksheet1.merge_cells(start_row=row,start_column=col,end_row=row+1,end_column=col)
            h8.alignment=Alignment(horizontal='center',vertical='center')
            h8.font=Font(size=12,bold=True)
            col+=1
            
            h9=worksheet1.cell(column=col, row=row, value="Name of Training\nIntervention/Pivotal\nProgramme Qualification")
            worksheet1.merge_cells(start_row=row,start_column=col,end_row=row+1,end_column=col)
            h9.alignment=Alignment(horizontal='center',vertical='center')
            h9.font=Font(size=12,bold=True)
            col+=1
            
            h10=worksheet1.cell(column=col, row=row, value="Piotal Programme\nInstitution")
            worksheet1.merge_cells(start_row=row,start_column=col,end_row=row+1,end_column=col)
            h10.alignment=Alignment(horizontal='center',vertical='center')
            h10.font=Font(size=12,bold=True)
            col+=1
            
            h11=worksheet1.cell(column=col, row=row, value="NQF\nAligned")
            worksheet1.merge_cells(start_row=row,start_column=col,end_row=row+1,end_column=col)
            h11.alignment=Alignment(horizontal='center',vertical='center')
            h11.font=Font(size=12,bold=True)
            col+=1
            
            h12=worksheet1.cell(column=col, row=row, value="NQF\nLevel")
            worksheet1.merge_cells(start_row=row,start_column=col,end_row=row+1,end_column=col)
            h12.alignment=Alignment(horizontal='center',vertical='center')
            h12.font=Font(size=12,bold=True)
            col+=1
            
            h13=worksheet1.cell(column=col, row=row, value="Cost")
            worksheet1.merge_cells(start_row=row,start_column=col,end_row=row+1,end_column=col)
            h13.alignment=Alignment(horizontal='center',vertical='center')
            h13.font=Font(size=12,bold=True)
            col+=1
            
            h14=worksheet1.cell(column=col, row=row, value="Start Date")
            worksheet1.merge_cells(start_row=row,start_column=col,end_row=row+1,end_column=col)
            h14.alignment=Alignment(horizontal='center',vertical='center')
            h14.font=Font(size=12,bold=True)
            col+=1
            
            h15=worksheet1.cell(column=col, row=row, value="End Date")
            worksheet1.merge_cells(start_row=row,start_column=col,end_row=row+1,end_column=col)
            h15.alignment=Alignment(horizontal='center',vertical='center')
            h15.font=Font(size=12,bold=True)
            col+=1
            
            h16=worksheet1.cell(column=col, row=row, value="AFRICAN")
            h16.alignment=Alignment(horizontal='center',vertical='center')
            h16.font=Font(size=12,bold=True)
            worksheet1.merge_cells(start_row=1,start_column=col,end_row=1,end_column=col+2)
            
            h16i=worksheet1.cell(column=col, row=row+1, value="M")
            h16i.font=Font(size=12,bold=True)
            
            col+=1
            h16j=worksheet1.cell(column=col, row=row+1, value="F")
            h16j.font=Font(size=12,bold=True)
            
            col+=1
            h16k=worksheet1.cell(column=col, row=row+1, value="D")
            h16k.font=Font(size=12,bold=True)
            
            col+=1
            h17=worksheet1.cell(column=col, row=row, value="COLOURED")
            h17.alignment=Alignment(horizontal='center',vertical='center')
            h17.font=Font(size=12,bold=True)
            worksheet1.merge_cells(start_row=1,start_column=col,end_row=1,end_column=col+2)
            
            h17i=worksheet1.cell(column=col, row=row+1, value="M")
            h17i.font=Font(size=12,bold=True)
            
            col+=1
            h17j=worksheet1.cell(column=col, row=row+1, value="F")
            h17j.font=Font(size=12,bold=True)
            
            col+=1
            h17k=worksheet1.cell(column=col, row=row+1, value="D")
            h17k.font=Font(size=12,bold=True)
            
            col+=1
            h18=worksheet1.cell(column=col, row=row, value="INDIAN")
            h18.alignment=Alignment(horizontal='center',vertical='center')
            h18.font=Font(size=12,bold=True)
            worksheet1.merge_cells(start_row=1,start_column=col,end_row=1,end_column=col+2)
            
            h18i=worksheet1.cell(column=col, row=row+1, value="M")
            h18i.font=Font(size=12,bold=True)
            
            col+=1
            h18j=worksheet1.cell(column=col, row=row+1, value="F")
            h18j.font=Font(size=12,bold=True)
            
            col+=1
            h18k=worksheet1.cell(column=col, row=row+1, value="D")
            h18k.font=Font(size=12,bold=True)
            
            
            col+=1
            h19=worksheet1.cell(column=col, row=row, value="WHITE")
            h19.alignment=Alignment(horizontal='center',vertical='center')
            h19.font=Font(size=12,bold=True)
            worksheet1.merge_cells(start_row=1,start_column=col,end_row=1,end_column=col+2)
            
            h19i=worksheet1.cell(column=col, row=row+1, value="M")
            h19i.font=Font(size=12,bold=True)
            
            col+=1
            h19j=worksheet1.cell(column=col, row=row+1, value="F")
            h19j.font=Font(size=12,bold=True)
            
            col+=1
            h19k=worksheet1.cell(column=col, row=row+1, value="D")
            h19k.font=Font(size=12,bold=True)
            
            col+=1
            h20=worksheet1.cell(column=col, row=row, value="Total")
            h20.alignment=Alignment(horizontal='center',vertical='center')
            h20.font=Font(size=12,bold=True)
            worksheet1.merge_cells(start_row=1,start_column=col,end_row=1,end_column=col+2)
            
            h20i=worksheet1.cell(column=col, row=row+1, value="M")
            h20i.font=Font(size=12,bold=True)
            
            col+=1
            h20j=worksheet1.cell(column=col, row=row+1, value="F")
            h20j.font=Font(size=12,bold=True)
            
            col+=1
            h20k=worksheet1.cell(column=col, row=row+1, value="D")
            h20k.font=Font(size=12,bold=True)
            
            col+=1
            h21=worksheet1.cell(column=col, row=row, value="AGE GROUP")
            h21.alignment=Alignment(horizontal='center',vertical='center')
            h21.font=Font(size=12,bold=True)
            worksheet1.merge_cells(start_row=1,start_column=col,end_row=1,end_column=col+2)
            
            h21i=worksheet1.cell(column=col, row=row+1, value="<35")
            h21i.font=Font(size=12,bold=True)
            
            col+=1
            h21j=worksheet1.cell(column=col, row=row+1, value="35-55")
            h21j.font=Font(size=12,bold=True)
            
            col+=1
            h21k=worksheet1.cell(column=col, row=row+1, value=">55")
            h21k.font=Font(size=12,bold=True)
            
            col+=1
            h22=worksheet1.cell(column=col, row=row, value="Total Cost")
            worksheet1.merge_cells(start_row=row,start_column=col,end_row=row+1,end_column=col)
            h22.alignment=Alignment(horizontal='center',vertical='center')
            h22.font=Font(size=12,bold=True)
            col+=1
            
            row=3
            col=1
            for actual_training_line in wsp_plan_data.training_actual_ids:
                row+=1
                col=1
                for actual_training_data in actual_training_line:

                    if actual_training_data.training_type != 0:
                        worksheet1.cell(column=col, row=row, value=self.get_training_type(actual_training_data.training_type))
                        col+=1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col+=1
                        
                    if actual_training_data.occupation:
                        worksheet1.cell(column=col, row=row, value=actual_training_data.occupation and actual_training_data.occupation.name)
                        col+=1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col+=1
                    if actual_training_data.specialization:
                        worksheet1.cell(column=col, row=row, value=actual_training_data.specialization and actual_training_data.specialization.name)
                        col+=1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col+=1
                    if actual_training_data.municipality_id:
                        worksheet1.cell(column=col, row=row, value=actual_training_data.municipality_id and actual_training_data.municipality_id.name)
                        col+=1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col+=1
                        
                    if actual_training_data.urban != 0:
                        worksheet1.cell(column=col, row=row, value=self.get_urban_rural(actual_training_data.urban))
                        col+=1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col+=1
                        
                        
                    if actual_training_data.learner_province:
                        worksheet1.cell(column=col, row=row, value=actual_training_data.learner_province and actual_training_data.learner_province.name)
                        col+=1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col+=1
                        
                    if actual_training_data.socio_economic_status !=0:
                        worksheet1.cell(column=col, row=row, value=self.get_socio_eco_status(actual_training_data.socio_economic_status))
                        col+=1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col+=1
                        
                    if actual_training_data.type_training:
                        worksheet1.cell(column=col, row=row, value=actual_training_data.type_training and actual_training_data.type_training.name)
                        col+=1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col+=1
                        
                    if actual_training_data.name_training !=0:
                        worksheet1.cell(column=col, row=row, value=actual_training_data.name_training)
                        col+=1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col+=1
                        
                    if actual_training_data.pivotal_programme_institution !=0:
                        worksheet1.cell(column=col, row=row, value=actual_training_data.pivotal_programme_institution)
                        col+=1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col+=1
                    if actual_training_data.nqf_aligned !=0:
                        worksheet1.cell(column=col, row=row, value=self.get_nqf_aligned(actual_training_data.nqf_aligned))
                        col+=1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col+=1
                    if actual_training_data.nqf_level !=0:
                        
                        worksheet1.cell(column=col, row=row, value=self.get_nqf_level(actual_training_data.nqf_level))
                        col+=1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col+=1
                        
                    if actual_training_data.training_cost !=0:
                        worksheet1.cell(column=col, row=row, value=actual_training_data.training_cost)
                        col+=1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col+=1
                    if  actual_training_data.start_date !=0:
                        worksheet1.cell(column=col, row=row, value=actual_training_data.start_date)
                        col+=1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col+=1
                        
                    if actual_training_data.end_date !=0:
                        worksheet1.cell(column=col, row=row, value=actual_training_data.end_date)
                        col+=1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col+=1
                    worksheet1.cell(column=col, row=row, value=actual_training_data.african_male)
                    col+=1
                    worksheet1.cell(column=col, row=row, value=actual_training_data.african_female)
                    col+=1
                    worksheet1.cell(column=col, row=row, value=actual_training_data.african_dissabled)
                    col+=1
                    worksheet1.cell(column=col, row=row, value=actual_training_data.coloured_male)
                    col+=1
                    worksheet1.cell(column=col, row=row, value=actual_training_data.coloured_female)
                    col+=1
                    worksheet1.cell(column=col, row=row, value=actual_training_data.coloured_dissabled)
                    col+=1
                    worksheet1.cell(column=col, row=row, value=actual_training_data.indian_male)
                    col+=1
                    worksheet1.cell(column=col, row=row, value=actual_training_data.indian_female)
                    col+=1
                    worksheet1.cell(column=col, row=row, value=actual_training_data.indian_dissabled)
                    col+=1
                    worksheet1.cell(column=col, row=row, value=actual_training_data.white_male)
                    col+=1
                    worksheet1.cell(column=col, row=row, value=actual_training_data.white_female)
                    col+=1
                    worksheet1.cell(column=col, row=row, value=actual_training_data.white_dissabled)
                    col+=1
                    worksheet1.cell(column=col, row=row, value=actual_training_data.total_male)
                    col+=1
                    worksheet1.cell(column=col, row=row, value=actual_training_data.total_female)
                    col+=1
                    worksheet1.cell(column=col, row=row, value=actual_training_data.total_dissabled)
                    col+=1
                    worksheet1.cell(column=col, row=row, value=actual_training_data.age_group_less)
                    col+=1
                    worksheet1.cell(column=col, row=row, value=actual_training_data.age_group_upto)
                    col+=1
                    worksheet1.cell(column=col, row=row, value=actual_training_data.age_group_greater)
                    col+=1
                    worksheet1.cell(column=col, row=row, value=actual_training_data.total_cost)
                    
        if wsp_plan_data.actual_adult_education_ids :
            row=col=1
            worksheet2 = workbook.create_sheet(title="Adult Education and Training")
            
            h1=worksheet2.cell(column=col, row=row, value="First Name")
            h1.alignment=Alignment(horizontal='center',vertical='center')
            h1.font=Font(size=12,bold=True)
            col+=1
            
            h2=worksheet2.cell(column=col, row=row, value="SurName")
            h2.alignment=Alignment(horizontal='center',vertical='center')
            h2.font=Font(size=12,bold=True)
            col+=1
            
            h3=worksheet2.cell(column=col, row=row, value="Id Number")
            h3.alignment=Alignment(horizontal='center',vertical='center')
            h3.font=Font(size=12,bold=True)
            col+=1
            
            h33=worksheet2.cell(column=col, row=row, value="Population Group")
            h33.alignment=Alignment(horizontal='center',vertical='center')
            h33.font=Font(size=12,bold=True)
            col+=1
            
            h4=worksheet2.cell(column=col, row=row, value="Gender")
            h4.alignment=Alignment(horizontal='center',vertical='center')
            h4.font=Font(size=12,bold=True)
            col+=1
            
            h5=worksheet2.cell(column=col, row=row, value="Disability Status And Type")
            h5.alignment=Alignment(horizontal='center',vertical='center')
            h5.font=Font(size=12,bold=True)
            col+=1
            
            h6=worksheet2.cell(column=col, row=row, value="Municipality")
            h6.alignment=Alignment(horizontal='center',vertical='center')
            h6.font=Font(size=12,bold=True)
            col+=1
            
            h7=worksheet2.cell(column=col, row=row, value="Learner Province")
            h7.alignment=Alignment(horizontal='center',vertical='center')
            h7.font=Font(size=12,bold=True)
            col+=1
            
            h8=worksheet2.cell(column=col, row=row, value="Urban/Rural")
            h8.alignment=Alignment(horizontal='center',vertical='center')
            h8.font=Font(size=12,bold=True)
            col+=1
            
            h9=worksheet2.cell(column=col, row=row, value="AET Start Date")
            h9.alignment=Alignment(horizontal='center',vertical='center')
            h9.font=Font(size=12,bold=True)
            col+=1
            
            h10=worksheet2.cell(column=col, row=row, value="AET End Date")
            h10.alignment=Alignment(horizontal='center',vertical='center')
            h10.font=Font(size=12,bold=True)
            col+=1
            
            h10=worksheet2.cell(column=col, row=row, value="Provider")
            h10.alignment=Alignment(horizontal='center',vertical='center')
            h10.font=Font(size=12,bold=True)
            col+=1
            
            h11=worksheet2.cell(column=col, row=row, value="AET Level")
            h11.alignment=Alignment(horizontal='center',vertical='center')
            h11.font=Font(size=12,bold=True)
            col+=1
            
            h12=worksheet2.cell(column=col, row=row, value="AET Subject")
            h12.alignment=Alignment(horizontal='center',vertical='center')
            h12.font=Font(size=12,bold=True)
            
            for actual_adult_line in wsp_plan_data.actual_adult_education_ids:
                row+=1
                col=1
                for actual_adult_data in actual_adult_line:
                    
                    if actual_adult_data.name != 0:
                        worksheet2.cell(column=col, row=row, value=actual_adult_data.name)
                        col+=1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col+=1
                        
                    if actual_adult_data.surname != 0:
                        worksheet2.cell(column=col, row=row, value=actual_adult_data.surname)
                        col+=1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col+=1
                        
                    if actual_adult_data.id_number != 0:
                        worksheet2.cell(column=col, row=row, value=actual_adult_data.id_number)
                        col+=1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col+=1
                    
                    if actual_adult_data.population_group != 0:
                        worksheet2.cell(column=col, row=row, value=self.get_population_group(actual_adult_data.population_group))
                        col+=1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col+=1
                    
                    if actual_adult_data.gender != 0:
                        worksheet2.cell(column=col, row=row, value=self.get_genders(actual_adult_data.gender))
                        col+=1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col+=1
                        
                    if actual_adult_data.dissability_status_and_type != 0:
                        worksheet2.cell(column=col, row=row, value=self.get_disability_status(actual_adult_data.dissability_status_and_type))
                        col+=1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col+=1
                    
                    if actual_adult_data.municipality_id:
                        worksheet2.cell(column=col, row=row, value=actual_adult_data.municipality_id and actual_adult_data.municipality_id.name)
                        col+=1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col+=1
                    
                    if actual_adult_data.province:
                        worksheet2.cell(column=col, row=row, value=actual_adult_data.province and actual_adult_data.province.name)
                        col+=1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col+=1

                    if actual_adult_data.urban !=0:
                        worksheet2.cell(column=col, row=row, value=self.get_urban_rural(actual_adult_data.urban))
                        col+=1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col+=1
                    
                    if actual_adult_data.start_date !=0:
                        worksheet2.cell(column=col, row=row, value=actual_adult_data.start_date)
                        col+=1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col+=1
                    
                    if actual_adult_data.end_date != 0:
                        worksheet2.cell(column=col, row=row, value=actual_adult_data.end_date)
                        col+=1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col+=1
                    
                    if actual_adult_data.provider != 0:
                        worksheet2.cell(column=col, row=row, value=actual_adult_data.provider)
                        col+=1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col+=1
                    
                    if actual_adult_data.aet_level != 0:
                        worksheet2.cell(column=col, row=row, value=self.get_aet_level(actual_adult_data.aet_level))
                        col+=1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col+=1
                    
                    if actual_adult_data.aet_subject != 0:
                        worksheet2.cell(column=col, row=row, value=self.get_aet_subject(actual_adult_data.aet_subject))
                        col+=1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col+=1
                    
        if wsp_plan_data.actual_adult_education_ids or wsp_plan_data.training_actual_ids :
            workbook.save(buffer)
            out_data = base64.encodestring(buffer.getvalue())
            attachment_obj = self.env['ir.attachment']
            new_attach = attachment_obj.create({
                'name':'ATR.xlsx',
                'res_name': 'wsp_data',
                'type': 'binary',
                'res_model': 'wsp.plan',
                'datas':out_data,
            })
            return {
            'type' : 'ir.actions.act_url',
            'url': '/web/binary/saveas?model=ir.attachment&field=datas&filename_field=name&id=%s' % ( new_attach.id, ),
            'target': 'self',
            }
        else:
            raise Warning(_('No data for Annual Training Report'))
            return True
          
    @api.multi
    def export_previous_pdf(self):
        data={}
        for rec in self:
            data['wsp_id'] = self._context.get('active_id',False)
            value = self.env['report'].get_action(self, 'hwseta_sdp.atr_report', data=data)
            return value   
            
    @api.multi
    def export_previous_wsp_pdf(self):
        data={}
        for rec in self:
            data['wsp_id'] = self._context.get('active_id',False)
            value = self.env['report'].get_action(self, 'hwseta_sdp.wsp_report', data=data)
            return value   
                    
    @api.multi
    def export_previous_wsp(self):
        ''' This Method will export wsp template. '''

        workbook=Workbook()
        buffer = cStringIO.StringIO()
        context = self._context
        wsp_plan_data = self.env['wsp.plan'].browse(context.get('active_id',False))
        
        
        if wsp_plan_data.total_employment_profile_ids :
            row=col=1
#                 worksheet1 = workbook.create_sheet(title="Actual Training")
            worksheet1=workbook.active
            worksheet1.title = "Total Employment Profile"
            
            h1=worksheet1.cell(column=col, row=row, value="SDL Number")
            worksheet1.merge_cells(start_row=row,start_column=col,end_row=row+1,end_column=col)
            h1.alignment=Alignment(horizontal='center',vertical='center')
            h1.font=Font(size=12,bold=True)
            col+=1
            
            h2=worksheet1.cell(column=col, row=row, value="First Name")
            worksheet1.merge_cells(start_row=row,start_column=col,end_row=row+1,end_column=col)
            h2.alignment=Alignment(horizontal='center',vertical='center')
            h2.font=Font(size=12,bold=True)
            col+=1
            
            h3=worksheet1.cell(column=col, row=row, value="Last Name")
            worksheet1.merge_cells(start_row=row,start_column=col,end_row=row+1,end_column=col)
            h3.alignment=Alignment(horizontal='center',vertical='center')
            h3.font=Font(size=12,bold=True)
            col+=1
            
            h4=worksheet1.cell(column=col, row=row, value="Citizen Status")
            worksheet1.merge_cells(start_row=row,start_column=col,end_row=row+1,end_column=col)
            h4.alignment=Alignment(horizontal='center',vertical='center')
            h4.font=Font(size=12,bold=True)
            col+=1
            
            h5=worksheet1.cell(column=col, row=row, value="Employee Id")
            worksheet1.merge_cells(start_row=row,start_column=col,end_row=row+1,end_column=col)
            h5.alignment=Alignment(horizontal='center',vertical='center')
            h5.font=Font(size=12,bold=True)
            col+=1
            
            h6=worksheet1.cell(column=col, row=row, value="ID Type")
            worksheet1.merge_cells(start_row=row,start_column=col,end_row=row+1,end_column=col)
            h6.alignment=Alignment(horizontal='center',vertical='center')
            h6.font=Font(size=12,bold=True)
            col+=1
            
            h7=worksheet1.cell(column=col, row=row, value="Date Of Birth")
            worksheet1.merge_cells(start_row=row,start_column=col,end_row=row+1,end_column=col)
            h7.alignment=Alignment(horizontal='center',vertical='center')
            h7.font=Font(size=12,bold=True)
            col+=1
            
            h8=worksheet1.cell(column=col, row=row, value="OFO Code")
            worksheet1.merge_cells(start_row=row,start_column=col,end_row=row+1,end_column=col)
            h8.alignment=Alignment(horizontal='center',vertical='center')
            h8.font=Font(size=12,bold=True)
            col+=1
            
            h9=worksheet1.cell(column=col, row=row, value="Occupation")
            worksheet1.merge_cells(start_row=row,start_column=col,end_row=row+1,end_column=col)
            h9.alignment=Alignment(horizontal='center',vertical='center')
            h9.font=Font(size=12,bold=True)
            col+=1
            
            h10=worksheet1.cell(column=col, row=row, value="Specialisation")
            worksheet1.merge_cells(start_row=row,start_column=col,end_row=row+1,end_column=col)
            h10.alignment=Alignment(horizontal='center',vertical='center')
            h10.font=Font(size=12,bold=True)
            col+=1
            
            h11=worksheet1.cell(column=col, row=row, value="City")
            worksheet1.merge_cells(start_row=row,start_column=col,end_row=row+1,end_column=col)
            h11.alignment=Alignment(horizontal='center',vertical='center')
            h11.font=Font(size=12,bold=True)
            col+=1
            
            h12=worksheet1.cell(column=col, row=row, value="Province")
            worksheet1.merge_cells(start_row=row,start_column=col,end_row=row+1,end_column=col)
            h12.alignment=Alignment(horizontal='center',vertical='center')
            h12.font=Font(size=12,bold=True)
            col+=1
            
            h13=worksheet1.cell(column=col, row=row, value="Urban/Rural")
            worksheet1.merge_cells(start_row=row,start_column=col,end_row=row+1,end_column=col)
            h13.alignment=Alignment(horizontal='center',vertical='center')
            h13.font=Font(size=12,bold=True)
            col+=1
            
            h14=worksheet1.cell(column=col, row=row, value="Highest Education Level")
            worksheet1.merge_cells(start_row=row,start_column=col,end_row=row+1,end_column=col)
            h14.alignment=Alignment(horizontal='center',vertical='center')
            h14.font=Font(size=12,bold=True)
            col+=1
            
            h15=worksheet1.cell(column=col, row=row, value="Race")
            worksheet1.merge_cells(start_row=row,start_column=col,end_row=row+1,end_column=col)
            h15.alignment=Alignment(horizontal='center',vertical='center')
            h15.font=Font(size=12,bold=True)
            col+=1
            
            h16=worksheet1.cell(column=col, row=row, value="Gender")
            worksheet1.merge_cells(start_row=row,start_column=col,end_row=row+1,end_column=col)
            h16.alignment=Alignment(horizontal='center',vertical='center')
            h16.font=Font(size=12,bold=True)
            col +=1
            
            h17=worksheet1.cell(column=col, row=row, value="Disability")
            worksheet1.merge_cells(start_row=row,start_column=col,end_row=row+1,end_column=col)
            h17.alignment=Alignment(horizontal='center',vertical='center')
            h17.font=Font(size=12,bold=True)
            
            row=3
            col=1
            for total_employment_profile_line in wsp_plan_data.total_employment_profile_ids:
                row+=1
                col=1
                for total_employment_profile_data in total_employment_profile_line:

                    if total_employment_profile_data.sdl_number:
                        worksheet1.cell(column=col, row=row, value=total_employment_profile_data.sdl_number and total_employment_profile_data.sdl_number.name)
                        col+=1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col+=1
                    
                    if total_employment_profile_data.name !=0:
                        worksheet1.cell(column=col, row=row, value=total_employment_profile_data.name)
                        col+=1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col+=1
                    
                    
                    if total_employment_profile_data.surname !=0:
                        worksheet1.cell(column=col, row=row, value=total_employment_profile_data.surname)
                        col+=1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col+=1
                    
                    if total_employment_profile_data.citizen_resident_status_code !=0:
                        worksheet1.cell(column=col, row=row, value=self.get_citizen_status(total_employment_profile_data.citizen_resident_status_code))
                        col+=1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col+=1
                    
                    if total_employment_profile_data.employee_id !=0:
                        worksheet1.cell(column=col, row=row, value=total_employment_profile_data.employee_id)
                        col+=1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col+=1
                    
                    if total_employment_profile_data.id_type !=0:
                        worksheet1.cell(column=col, row=row, value=self.get_id_type(total_employment_profile_data.id_type))
                        col+=1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col+=1
                        
                    if total_employment_profile_data.dob !=0:
                        worksheet1.cell(column=col, row=row, value=total_employment_profile_data.dob)
                        col+=1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col+=1
                    
                    if total_employment_profile_data.ofo_code:
                        worksheet1.cell(column=col, row=row, value=total_employment_profile_data.ofo_code and total_employment_profile_data.ofo_code.name)
                        col+=1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col+=1
                        
                    if total_employment_profile_data.occupation:
                        worksheet1.cell(column=col, row=row, value=total_employment_profile_data.occupation and total_employment_profile_data.occupation.name)
                        col+=1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col+=1
                        
                    if total_employment_profile_data.specialization:
                        worksheet1.cell(column=col, row=row, value=total_employment_profile_data.ofo_code and total_employment_profile_data.specialization.name)
                        col+=1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col+=1
                        
                    if total_employment_profile_data.city_id:
                        worksheet1.cell(column=col, row=row, value=total_employment_profile_data.city_id and total_employment_profile_data.city_id.name)
                        col+=1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col+=1
                        
                    if total_employment_profile_data.province:
                        worksheet1.cell(column=col, row=row, value=total_employment_profile_data.province and total_employment_profile_data.province.name)
                        col+=1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col+=1
                        
                    if total_employment_profile_data.urban != 0:
                        worksheet1.cell(column=col, row=row, value=self.get_urban_rural(total_employment_profile_data.urban))
                        col+=1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col+=1
                    
                    if total_employment_profile_data.highest_education_level != 0:
                        worksheet1.cell(column=col, row=row, value=self.get_highest_edu(total_employment_profile_data.highest_education_level))
                        col+=1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col+=1
                        
                    if total_employment_profile_data.population_group != 0:
                        worksheet1.cell(column=col, row=row, value=self.get_population_group(total_employment_profile_data.population_group))
                        col+=1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col+=1
                        
                    if total_employment_profile_data.gender != 0:
                        worksheet1.cell(column=col, row=row, value=self.get_genders(total_employment_profile_data.gender))
                        col+=1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col+=1
                        
                    if total_employment_profile_data.dissability != 0:
                        worksheet1.cell(column=col, row=row, value=self.get_nqf_aligned(total_employment_profile_data.dissability))
                        col+=1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col+=1
                        
            
        if wsp_plan_data.training_planned_ids :
            row=col=1
            worksheet2 = workbook.create_sheet(title="Planned Training")
            
            h1=worksheet2.cell(column=col, row=row, value="Type Of Training")
            h1.alignment=Alignment(horizontal='center',vertical='center')
            h1.font=Font(size=12,bold=True)
            col+=1
            
            h2=worksheet2.cell(column=col, row=row, value="Name")
            h2.alignment=Alignment(horizontal='center',vertical='center')
            h2.font=Font(size=12,bold=True)
            col+=1
            
            h3=worksheet2.cell(column=col, row=row, value="SurName")
            h3.alignment=Alignment(horizontal='center',vertical='center')
            h3.font=Font(size=12,bold=True)
            col+=1
            
            h4=worksheet2.cell(column=col, row=row, value="Employee Id")
            h4.alignment=Alignment(horizontal='center',vertical='center')
            h4.font=Font(size=12,bold=True)
            col+=1
            
            h5=worksheet2.cell(column=col, row=row, value="OFO Code")
            h5.alignment=Alignment(horizontal='center',vertical='center')
            h5.font=Font(size=12,bold=True)
            col+=1
            
            h6=worksheet2.cell(column=col, row=row, value="Occupation")
            h6.alignment=Alignment(horizontal='center',vertical='center')
            h6.font=Font(size=12,bold=True)
            col+=1
            
            h7=worksheet2.cell(column=col, row=row, value="Specialisation")
            h7.alignment=Alignment(horizontal='center',vertical='center')
            h7.font=Font(size=12,bold=True)
            col+=1
            
            h8=worksheet2.cell(column=col, row=row, value="City")
            h8.alignment=Alignment(horizontal='center',vertical='center')
            h8.font=Font(size=12,bold=True)
            col+=1
            
            h9=worksheet2.cell(column=col, row=row, value="Urban/Rural")
            h9.alignment=Alignment(horizontal='center',vertical='center')
            h9.font=Font(size=12,bold=True)
            col+=1
            
            h10=worksheet2.cell(column=col, row=row, value="Province")
            h10.alignment=Alignment(horizontal='center',vertical='center')
            h10.font=Font(size=12,bold=True)
            col+=1
            
            h11=worksheet2.cell(column=col, row=row, value="Employed/UnEmployed")
            h11.alignment=Alignment(horizontal='center',vertical='center')
            h11.font=Font(size=12,bold=True)
            col+=1
            
            h12=worksheet2.cell(column=col, row=row, value="Type of Training Intervention")
            h12.alignment=Alignment(horizontal='center',vertical='center')
            h12.font=Font(size=12,bold=True)
            col+=1
            
            h13=worksheet2.cell(column=col, row=row, value="Other Type Of Training Intervention/\nPivotal programme Type")
            h13.alignment=Alignment(horizontal='center',vertical='center')
            h13.font=Font(size=12,bold=True)
            col+=1
            
            h14=worksheet2.cell(column=col, row=row, value="Name of training\nIntervention")
            h14.alignment=Alignment(horizontal='center',vertical='center')
            h14.font=Font(size=12,bold=True)
            col+=1
            
                         
            h16=worksheet2.cell(column=col, row=row, value="Pivotal Programme\nQualification")
            h16.alignment=Alignment(horizontal='center',vertical='center')
            h16.font=Font(size=12,bold=True)
            col+=1
            
            h17=worksheet2.cell(column=col, row=row, value="Pivotal Programme\ninstitution")
            h17.alignment=Alignment(horizontal='center',vertical='center')
            h17.font=Font(size=12,bold=True)
            col+=1
            
            h18=worksheet2.cell(column=col, row=row, value="Cost Per Learner")
            h18.alignment=Alignment(horizontal='center',vertical='center')
            h18.font=Font(size=12,bold=True)
            col+=1
            
            h19=worksheet2.cell(column=col, row=row, value="Start Date")
            h19.alignment=Alignment(horizontal='center',vertical='center')
            h19.font=Font(size=12,bold=True)
            col+=1
            
            h20=worksheet2.cell(column=col, row=row, value="End Date")
            h20.alignment=Alignment(horizontal='center',vertical='center')
            h20.font=Font(size=12,bold=True)
            col+=1
            
            h21=worksheet2.cell(column=col, row=row, value="NQF\nAligned")
            h21.alignment=Alignment(horizontal='center',vertical='center')
            h21.font=Font(size=12,bold=True)
            col+=1
            
            h22=worksheet2.cell(column=col, row=row, value="NQF Level")
            h22.alignment=Alignment(horizontal='center',vertical='center')
            h22.font=Font(size=12,bold=True)
            col+=1
            
            h23=worksheet2.cell(column=col, row=row, value="Race")
            h23.alignment=Alignment(horizontal='center',vertical='center')
            h23.font=Font(size=12,bold=True)
            col+=1
            
            h24=worksheet2.cell(column=col, row=row, value="Gender")
            h24.alignment=Alignment(horizontal='center',vertical='center')
            h24.font=Font(size=12,bold=True)
            col+=1
            
            h25=worksheet2.cell(column=col, row=row, value="Disability")
            h25.alignment=Alignment(horizontal='center',vertical='center')
            h25.font=Font(size=12,bold=True)
            col+=1
            
            
            for training_planned_line in wsp_plan_data.training_planned_ids:
                row+=1
                col=1
                for training_planned_data in training_planned_line:
                    
                    if training_planned_data.training_type != 0:
                        worksheet2.cell(column=col, row=row, value=self.get_training_type(training_planned_data.training_type))
                        col+=1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col+=1
                        
                    if training_planned_data.name != 0:
                        worksheet2.cell(column=col, row=row, value=training_planned_data.name)
                        col+=1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col+=1
                        
                    if training_planned_data.surname != 0:
                        worksheet2.cell(column=col, row=row, value=training_planned_data.surname)
                        col+=1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col+=1
                        
                    if training_planned_data.employee_id != 0:
                        worksheet2.cell(column=col, row=row, value=training_planned_data.employee_id)
                        col+=1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col+=1   
                        
                    if training_planned_data.code:
                        worksheet2.cell(column=col, row=row, value=training_planned_data.code and training_planned_data.code.name)
                        col+=1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col+=1
                    
                    if training_planned_data.occupation:
                        worksheet2.cell(column=col, row=row, value=training_planned_data.occupation and training_planned_data.occupation.name)
                        col+=1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col+=1
                        
                    if training_planned_data.specialization:
                        worksheet2.cell(column=col, row=row, value=training_planned_data.specialization and training_planned_data.specialization.name)
                        col+=1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col+=1
                        
                    if training_planned_data.city_id:
                        worksheet2.cell(column=col, row=row, value=training_planned_data.city_id and training_planned_data.city_id.name)
                        col+=1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col+=1
                        
                    if training_planned_data.urban != 0:
                        worksheet2.cell(column=col, row=row, value=self.get_urban_rural(training_planned_data.urban))
                        col+=1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col+=1
                            
                    if training_planned_data.learner_province:
                        worksheet2.cell(column=col, row=row, value=training_planned_data.learner_province and training_planned_data.learner_province.name)
                        col+=1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col+=1
                        
                    if training_planned_data.socio_economic_status != 0:
                        worksheet2.cell(column=col, row=row, value=self.get_socio_eco_status(training_planned_data.socio_economic_status))
                        col+=1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col+=1
                        
                    if training_planned_data.type_training:
                        worksheet2.cell(column=col, row=row, value=training_planned_data.type_training and training_planned_data.type_training.name)
                        col+=1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col+=1
                    
                    if training_planned_data.other_type_of_intervention != 0:
                        worksheet2.cell(column=col, row=row, value=training_planned_data.other_type_of_intervention)
                        col+=1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col+=1
                    
                    if training_planned_data.name_training != 0:
                        worksheet2.cell(column=col, row=row, value=training_planned_data.name_training)
                        col+=1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col+=1
                    
                    if training_planned_data.pivotal_programme_qualification != 0:
                        worksheet2.cell(column=col, row=row, value=training_planned_data.pivotal_programme_qualification)
                        col+=1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col+=1
                    
                    if training_planned_data.pivotal_programme_institution != 0:
                        worksheet2.cell(column=col, row=row, value=training_planned_data.pivotal_programme_institution)
                        col+=1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col+=1
                    
                    if training_planned_data.training_cost != 0:
                        worksheet2.cell(column=col, row=row, value=training_planned_data.training_cost)
                        col+=1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col+=1
                    
                    if training_planned_data.start_date != 0:
                        worksheet2.cell(column=col, row=row, value=training_planned_data.start_date)
                        col+=1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col+=1
                        
                    if training_planned_data.end_date != 0:
                        worksheet2.cell(column=col, row=row, value=training_planned_data.end_date)
                        col+=1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col+=1
                        
                    if training_planned_data.nqf_aligned != 0:
                        worksheet2.cell(column=col, row=row, value=self.get_nqf_aligned(training_planned_data.nqf_aligned))
                        col+=1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col+=1
                        
                    if training_planned_data.nqf_level != 0:
                        worksheet2.cell(column=col, row=row, value=self.get_nqf_level(training_planned_data.nqf_level))
                        col+=1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col+=1
                        
                    if training_planned_data.population_group != 0:
                        worksheet2.cell(column=col, row=row, value=self.get_population_group(training_planned_data.population_group))
                        col+=1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col+=1
                        
                    if training_planned_data.gender != 0:
                        worksheet2.cell(column=col, row=row, value=self.get_genders(training_planned_data.gender))
                        col+=1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col+=1
                        
                    if training_planned_data.dissability != 0:
                        worksheet2.cell(column=col, row=row, value=self.get_nqf_aligned(training_planned_data.dissability))
                        col+=1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col+=1
        
        
        if wsp_plan_data.planned_adult_education_ids :
            row=col=1
            worksheet3 = workbook.create_sheet(title="Adult Education and Training")
            
            h1=worksheet3.cell(column=col, row=row, value="First Name")
            h1.alignment=Alignment(horizontal='center',vertical='center')
            h1.font=Font(size=12,bold=True)
            col+=1
            
            h2=worksheet3.cell(column=col, row=row, value="Sur Name")
            h2.alignment=Alignment(horizontal='center',vertical='center')
            h2.font=Font(size=12,bold=True)
            col+=1
            
                          
            h4=worksheet3.cell(column=col, row=row, value="ID Number")
            h4.alignment=Alignment(horizontal='center',vertical='center')
            h4.font=Font(size=12,bold=True)
            col+=1
            
            h5=worksheet3.cell(column=col, row=row, value="Population Group")
            h5.alignment=Alignment(horizontal='center',vertical='center')
            h5.font=Font(size=12,bold=True)
            col+=1
            
            h6=worksheet3.cell(column=col, row=row, value="Gender")
            h6.alignment=Alignment(horizontal='center',vertical='center')
            h6.font=Font(size=12,bold=True)
            col+=1
            
            h7=worksheet3.cell(column=col, row=row, value="Disability Status And Type")
            h7.alignment=Alignment(horizontal='center',vertical='center')
            h7.font=Font(size=12,bold=True)
            col+=1
            
            h8=worksheet3.cell(column=col, row=row, value="City")
            h8.alignment=Alignment(horizontal='center',vertical='center')
            h8.font=Font(size=12,bold=True)
            col+=1
            
            h9=worksheet3.cell(column=col, row=row, value="Learner Province")
            h9.alignment=Alignment(horizontal='center',vertical='center')
            h9.font=Font(size=12,bold=True)
            col+=1
            
            h10=worksheet3.cell(column=col, row=row, value="Urban/Rural")
            h10.alignment=Alignment(horizontal='center',vertical='center')
            h10.font=Font(size=12,bold=True)
            col+=1
            
            h11=worksheet3.cell(column=col, row=row, value="AET Start Date")
            h11.alignment=Alignment(horizontal='center',vertical='center')
            h11.font=Font(size=12,bold=True)
            col+=1
            
            h12=worksheet3.cell(column=col, row=row, value="AET End Date")
            h12.alignment=Alignment(horizontal='center',vertical='center')
            h12.font=Font(size=12,bold=True)
            col+=1
            
            h13=worksheet3.cell(column=col, row=row, value="Provider")
            h13.alignment=Alignment(horizontal='center',vertical='center')
            h13.font=Font(size=12,bold=True)
            col+=1
            
            h14=worksheet3.cell(column=col, row=row, value="AET Level")
            h14.alignment=Alignment(horizontal='center',vertical='center')
            h14.font=Font(size=12,bold=True)
            col+=1
            
                         
            h15=worksheet3.cell(column=col, row=row, value="AET Subject")
            h15.alignment=Alignment(horizontal='center',vertical='center')
            h15.font=Font(size=12,bold=True)
            col+=1
           
            
            
            for planned_adult_education_line in wsp_plan_data.planned_adult_education_ids:
                row+=1
                col=1
                for planned_adult_education_data in planned_adult_education_line:
                    
                    if planned_adult_education_data.name != 0:
                        worksheet3.cell(column=col, row=row, value=planned_adult_education_data.name)
                        col+=1
                    else:
                        worksheet3.cell(column=col, row=row, value='')
                        col+=1
                    
                    if planned_adult_education_data.surname != 0:
                        worksheet3.cell(column=col, row=row, value=planned_adult_education_data.surname)
                        col+=1
                    else:
                        worksheet3.cell(column=col, row=row, value='')
                        col+=1
                        
                    if planned_adult_education_data.id_number != 0:
                        worksheet3.cell(column=col, row=row, value=planned_adult_education_data.id_number)
                        col+=1
                    else:
                        worksheet3.cell(column=col, row=row, value='')
                        col+=1   
                    
                    if planned_adult_education_data.population_group != 0:
                        worksheet3.cell(column=col, row=row, value=self.get_population_group(planned_adult_education_data.population_group))
                        col+=1
                    else:
                        worksheet3.cell(column=col, row=row, value='')
                        col+=1
                        
                    if planned_adult_education_data.gender != 0:
                        worksheet3.cell(column=col, row=row, value=self.get_genders(planned_adult_education_data.gender))
                        col+=1
                    else:
                        worksheet3.cell(column=col, row=row, value='')
                        col+=1
                    
                    if planned_adult_education_data.dissability_status_and_type != 0:
                        worksheet3.cell(column=col, row=row, value=self.get_disability_status(planned_adult_education_data.dissability_status_and_type))
                        col+=1
                    else:
                        worksheet3.cell(column=col, row=row, value='')
                        col+=1
                    
                    if planned_adult_education_data.city_id:
                        worksheet3.cell(column=col, row=row, value=planned_adult_education_data.city_id and planned_adult_education_data.city_id.name)
                        col+=1
                    else:
                        worksheet3.cell(column=col, row=row, value='')
                        col+=1
                    
                    if planned_adult_education_data.province:
                        worksheet3.cell(column=col, row=row, value=planned_adult_education_data.province and planned_adult_education_data.province.name)
                        col+=1
                    else:
                        worksheet3.cell(column=col, row=row, value='')
                        col+=1
                    
                    if planned_adult_education_data.urban != 0:
                        worksheet3.cell(column=col, row=row, value=self.get_urban_rural(planned_adult_education_data.urban))
                        col+=1
                    else:
                        worksheet3.cell(column=col, row=row, value='')
                        col+=1
                    
                    if planned_adult_education_data.start_date != 0:
                        worksheet3.cell(column=col, row=row, value=planned_adult_education_data.start_date)
                        col+=1
                    else:
                        worksheet3.cell(column=col, row=row, value='')
                        col+=1
                        
                    if planned_adult_education_data.end_date != 0:
                        worksheet3.cell(column=col, row=row, value=planned_adult_education_data.end_date)
                        col+=1
                    else:
                        worksheet3.cell(column=col, row=row, value='')
                        col+=1
                    
                    if planned_adult_education_data.provider != 0:
                        worksheet3.cell(column=col, row=row, value=planned_adult_education_data.provider)
                        col+=1
                    else:
                        worksheet3.cell(column=col, row=row, value='')
                        col+=1
                    
                    
                    if planned_adult_education_data.aet_level != 0:
                        worksheet3.cell(column=col, row=row, value=self.get_aet_level(planned_adult_education_data.aet_level))
                        col+=1
                    else:
                        worksheet3.cell(column=col, row=row, value='')
                        col+=1
                        
                    if planned_adult_education_data.aet_subject != 0:
                        worksheet3.cell(column=col, row=row, value=self.get_aet_subject(planned_adult_education_data.aet_subject))
                        col+=1
                    else:
                        worksheet3.cell(column=col, row=row, value='')
                        col+=1
                    
         
         
        if wsp_plan_data.scarce_and_critical_skills_ids :
            row=col=1
            worksheet4 = workbook.create_sheet(title="Vacancies Hard to Fill")
            
            h1=worksheet4.cell(column=col, row=row, value="OFO Code")
            h1.alignment=Alignment(horizontal='center',vertical='center')
            h1.font=Font(size=12,bold=True)
            col+=1
            
            h2=worksheet4.cell(column=col, row=row, value="Occupation")
            h2.alignment=Alignment(horizontal='center',vertical='center')
            h2.font=Font(size=12,bold=True)
            col+=1
            
                          
            h4=worksheet4.cell(column=col, row=row, value="Specialisation")
            h4.alignment=Alignment(horizontal='center',vertical='center')
            h4.font=Font(size=12,bold=True)
            col+=1
            
            h5=worksheet4.cell(column=col, row=row, value="Province")
            h5.alignment=Alignment(horizontal='center',vertical='center')
            h5.font=Font(size=12,bold=True)
            col+=1
            
            h6=worksheet4.cell(column=col, row=row, value="Number of Vacancies")
            h6.alignment=Alignment(horizontal='center',vertical='center')
            h6.font=Font(size=12,bold=True)
            col+=1
            
            h7=worksheet4.cell(column=col, row=row, value="Gender")
            h7.alignment=Alignment(horizontal='center',vertical='center')
            h7.font=Font(size=12,bold=True)
            col+=1
            
            h8=worksheet4.cell(column=col, row=row, value="Race")
            h8.alignment=Alignment(horizontal='center',vertical='center')
            h8.font=Font(size=12,bold=True)
            col+=1
            
            h9=worksheet4.cell(column=col, row=row, value="Number of months\nposition has")
            h9.alignment=Alignment(horizontal='center',vertical='center')
            h9.font=Font(size=12,bold=True)
            col+=1
            
            h10=worksheet4.cell(column=col, row=row, value="Comments")
            h10.alignment=Alignment(horizontal='center',vertical='center')
            h10.font=Font(size=12,bold=True)
            
            for scarce_and_critical_skills_line in wsp_plan_data.scarce_and_critical_skills_ids:
                row+=1
                col=1
                for scarce_and_critical_skills_data in scarce_and_critical_skills_line:
                    
                    if scarce_and_critical_skills_data.ofo_code:
                        worksheet4.cell(column=col, row=row, value=scarce_and_critical_skills_data.ofo_code and scarce_and_critical_skills_data.ofo_code.name)
                        col+=1
                    else:
                        worksheet4.cell(column=col, row=row, value='')
                        col+=1
                        
                    if scarce_and_critical_skills_data.occupation:
                        worksheet4.cell(column=col, row=row, value=scarce_and_critical_skills_data.occupation and scarce_and_critical_skills_data.occupation.name)
                        col+=1
                    else:
                        worksheet4.cell(column=col, row=row, value='')
                        col+=1
                        
                    if scarce_and_critical_skills_data.specialization:
                        worksheet4.cell(column=col, row=row, value=scarce_and_critical_skills_data.specialization and scarce_and_critical_skills_data.specialization.name)
                        col+=1
                    else:
                        worksheet4.cell(column=col, row=row, value='')
                        col+=1
                        
                    if scarce_and_critical_skills_data.province:
                        worksheet4.cell(column=col, row=row, value=scarce_and_critical_skills_data.province and scarce_and_critical_skills_data.province.name)
                        col+=1
                    else:
                        worksheet4.cell(column=col, row=row, value='')
                        col+=1
                        
                    if scarce_and_critical_skills_data.number_of_vacancies != 0:
                        worksheet4.cell(column=col, row=row, value=scarce_and_critical_skills_data.number_of_vacancies)
                        col+=1
                    else:
                        worksheet4.cell(column=col, row=row, value='')
                        col+=1  
                    
                    if scarce_and_critical_skills_data.gender != 0:
                        worksheet4.cell(column=col, row=row, value=self.get_genders(scarce_and_critical_skills_data.gender))
                        col+=1
                    else:
                        worksheet4.cell(column=col, row=row, value='')
                        col+=1
                    
                    if scarce_and_critical_skills_data.population_group != 0:
                        worksheet4.cell(column=col, row=row, value=self.get_population_group(scarce_and_critical_skills_data.population_group))
                        col+=1
                    else:
                        worksheet4.cell(column=col, row=row, value='')
                        col+=1
                    
                    if scarce_and_critical_skills_data.no_of_months != 0:
                        worksheet4.cell(column=col, row=row, value=scarce_and_critical_skills_data.no_of_months)
                        col+=1
                    else:
                        worksheet4.cell(column=col, row=row, value='')
                        col+=1
                    
                    if scarce_and_critical_skills_data.comments != 0:
                        worksheet4.cell(column=col, row=row, value=scarce_and_critical_skills_data.comments)
                        col+=1
                    else:
                        worksheet4.cell(column=col, row=row, value='')
                        col+=1
                    
        
        if wsp_plan_data.total_employment_profile_ids or wsp_plan_data.training_planned_ids or wsp_plan_data.planned_adult_education_ids or wsp_plan_data.scarce_and_critical_skills_ids:
        
            workbook.save(buffer)
            out_data = base64.encodestring(buffer.getvalue())    
            
            attachment_obj = self.env['ir.attachment']
            new_attach = attachment_obj.create({
                'name':'WSP.xlsx',
                'res_name': 'wsp_data',
                'type': 'binary',
                'res_model': 'wsp.plan',
                'datas': out_data,
            })
    
            return {
            'type' : 'ir.actions.act_url',
            'url': '/web/binary/saveas?model=ir.attachment&field=datas&filename_field=name&id=%s' % ( new_attach.id, ),
            'target': 'self',
            }
        else:
            raise Warning(_('No data for Workplace Skills Plan'))
            return True


    
export_wsp_xls()
